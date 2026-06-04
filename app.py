import streamlit as st
import re
import io
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="USSC Contract Extractor",
    page_icon="📋",
    layout="centered"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .main { background-color: #f8f9fb; }
    .block-container { padding-top: 2.5rem; max-width: 780px; }
    h1 { font-size: 1.8rem !important; font-weight: 600 !important; color: #1a2744 !important; margin-bottom: 0.25rem !important; }
    .subtitle { color: #6b7280; font-size: 0.95rem; margin-bottom: 2rem; }
    .info-box { background: #eef2ff; border-left: 4px solid #1a2744; padding: 0.85rem 1.1rem; border-radius: 0 8px 8px 0; font-size: 0.88rem; color: #374151; margin-bottom: 1.5rem; }
    .result-box { background: #f0fdf4; border-left: 4px solid #16a34a; padding: 0.85rem 1.1rem; border-radius: 0 8px 8px 0; font-size: 0.88rem; color: #166534; margin: 1rem 0; }
    .stButton > button { background-color: #1a2744; color: white; border: none; border-radius: 8px; padding: 0.6rem 1.8rem; font-family: 'DM Sans', sans-serif; font-weight: 500; font-size: 0.95rem; width: 100%; }
    .stButton > button:hover { background-color: #243660; }
    .stDownloadButton > button { background-color: #16a34a; color: white; border: none; border-radius: 8px; padding: 0.6rem 1.8rem; font-family: 'DM Sans', sans-serif; font-weight: 500; font-size: 0.95rem; width: 100%; }
    .stDownloadButton > button:hover { background-color: #15803d; }
    .file-count { font-size: 0.85rem; color: #6b7280; margin-top: 0.4rem; }
    footer {visibility: hidden;} #MainMenu {visibility: hidden;} header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Known USSC signers — add more as needed
# ─────────────────────────────────────────────
USSC_SIGNERS = [
    "Nora Osei", "Olivia Bowman", "Sarah Hebberd", "Tim Phelan",
    "Terrence Trammell", "Luke Gromer", "William Phelan",
    "Andrew Harrington", "Jessica Warren"
]

# ─────────────────────────────────────────────
# Text helpers
# ─────────────────────────────────────────────
def clean(text):
    if not text:
        return ""
    text = re.sub(r'_+', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def collapse_spaced(text):
    if not text:
        return text
    text = re.sub(r'_', '', text)
    def collapse(m):
        return m.group(0).replace(' ', '')
    text = re.sub(r'(?<!\w)([A-Za-z0-9] )+[A-Za-z0-9](?!\w)', collapse, text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def extract_all_dates(text):
    collapsed = collapse_spaced(text)
    # Only return values that look like real dates (not SSNs or phone numbers)
    candidates = re.findall(r'\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}', collapsed)
    valid = []
    for c in candidates:
        parts = re.split(r'[\/\-]', c)
        # A real date: month 1-12, day 1-31, year 2-4 digits
        if len(parts) == 3:
            try:
                month = int(parts[0])
                day = int(parts[1])
                year = int(parts[2])
                if 1 <= month <= 12 and 1 <= day <= 31 and (year > 100 or year < 100):
                    valid.append(c)
            except:
                pass
    return valid

# ─────────────────────────────────────────────
# Contract type detection
# ─────────────────────────────────────────────
def detect_contract_type(pages_text, full_text):
    """
    Entity: has insurance paragraph with Commercial General Liability / $1,000,000
    Individual with Entity Payment: no insurance paragraph, but signature page
        has two operator name lines (person name + LLC name) and page 1 mentions
        both a person AND an LLC/Entity
    Individual: no insurance paragraph, single operator name
    """
    # Check for Entity: insurance paragraph
    insurance_signals = [
        "Commercial General Liability",
        "1,000,000",
        "additional insureds",
        "employer's liability insurance",
        "business automobile liability"
    ]
    insurance_count = sum(1 for s in insurance_signals if s.lower() in full_text.lower())
    if insurance_count >= 2:
        return "Entity"

    # Check for Individual with Entity Payment:
    # Page 1 mentions both a person name and an LLC/Entity name in opening paragraph
    page1 = pages_text[0] if pages_text else ""
    
    # Look for "Entity" keyword in opening paragraph
    entity_signals = [
        '"Entity"',
        '("Entity")',
        'collectively, "Operator"',
        'collectively "Operator"',
    ]
    has_entity_keyword = any(s.lower() in full_text[:3000].lower() for s in entity_signals)
    
    # Also check signature page for two operator name lines
    sig_text = find_signature_page(pages_text)
    collapsed_sig = collapse_spaced(sig_text)
    
    # Look for LLC in operator block (before U.S. Sports Camps)
    ussc_pos = sig_text.find('U.S. Sports Camps') if sig_text else -1
    operator_block = sig_text[:ussc_pos] if ussc_pos > 0 else sig_text
    has_llc_in_sig = bool(re.search(r'\bLLC\b|\bInc\b|\bCorp\b|\bAthletics\b', operator_block, re.IGNORECASE))
    
    if has_entity_keyword or has_llc_in_sig:
        return "Individual with Entity Payment"

    return "Individual"

# ─────────────────────────────────────────────
# Header name extraction
# ─────────────────────────────────────────────
def extract_header_name(pages):
    if not pages:
        return "NOT FOUND"
    lines = [l.strip() for l in pages[0].split('\n') if l.strip()]
    year_pattern = re.compile(r'^20\d{2}$')
    for i, line in enumerate(lines):
        if year_pattern.match(line):
            if i > 0:
                candidate = lines[i-1]
                if 'DocuSign' not in candidate and 'Envelope' not in candidate and 'Docusign' not in candidate:
                    return clean(candidate)
    # fallback: skip DocuSign lines and return first real line
    for line in lines:
        if 'DocuSign' not in line and 'Envelope' not in line and 'Docusign' not in line:
            return clean(line)
    return "NOT FOUND"

# ─────────────────────────────────────────────
# Expiration date
# ─────────────────────────────────────────────
def extract_expiration_date(full_text):
    m = re.search(r'shall continue until\s+(December 31,\s*20\d{2})', full_text, re.IGNORECASE)
    if m:
        return clean(m.group(1))
    return "NOT FOUND"

# ─────────────────────────────────────────────
# Find signature page
# ─────────────────────────────────────────────
def find_signature_page(pages):
    for text in pages:
        if text and 'EXECUTED by the parties' in text:
            return text
    for text in pages:
        if text and 'U.S. Sports Camps' in text and 'Date' in text:
            return text
    return ""

# ─────────────────────────────────────────────
# Signature field extraction
# ─────────────────────────────────────────────
def extract_signature_fields(pages):
    operator_date = "NOT FOUND"
    ussc_name = "NOT FOUND"
    ussc_date = "NOT FOUND"

    sig_text = find_signature_page(pages)
    if not sig_text:
        return operator_date, ussc_name, ussc_date

    collapsed_full = collapse_spaced(sig_text)
    lines = sig_text.split('\n')
    collapsed_lines = [collapse_spaced(l) for l in lines]

    # ── USSC signer: check known signers first ──
    for signer in USSC_SIGNERS:
        if signer in collapsed_full:
            ussc_name = signer
            break

    # Fallback: last Name: occurrence after collapsing
    if ussc_name == "NOT FOUND":
        parts = re.split(r'Name:', collapsed_full)
        if len(parts) >= 2:
            last_part = parts[-1].strip()
            words = last_part.split()
            name_words = []
            for w in words:
                if re.match(r'^[A-Za-zÀ-ÿ\-]+$', w):
                    name_words.append(w)
                else:
                    break
                if len(name_words) == 2:
                    break
            if name_words and len(' '.join(name_words)) > 3:
                candidate = ' '.join(name_words)
                # Make sure it's not a label word
                skip_words = ['Director', 'Senior', 'Growth', 'Running', 'Sports', 'Camps']
                if not any(s == candidate for s in skip_words):
                    ussc_name = candidate

    # ── Dates ──
    for collapsed in collapsed_lines:
        # Lines with Date: label
        if re.search(r'\bDate:', collapsed):
            dates = extract_all_dates(collapsed)
            if len(dates) >= 2:
                if operator_date == "NOT FOUND":
                    operator_date = dates[0]
                if ussc_date == "NOT FOUND":
                    ussc_date = dates[1]
            elif len(dates) == 1:
                if operator_date == "NOT FOUND":
                    operator_date = dates[0]
                elif ussc_date == "NOT FOUND":
                    ussc_date = dates[0]

        # Lines with two dates but no Date: label (some contract formats)
        elif operator_date == "NOT FOUND" or ussc_date == "NOT FOUND":
            dates = extract_all_dates(collapsed)
            if len(dates) >= 2:
                if operator_date == "NOT FOUND":
                    operator_date = dates[0]
                if ussc_date == "NOT FOUND":
                    ussc_date = dates[1]
            elif len(dates) == 1:
                # Only assign standalone date if it looks like a signing date
                # (year between 2018-2030, not a phone/SSN)
                d = dates[0]
                parts = re.split(r'[\/\-]', d)
                if len(parts) == 3:
                    try:
                        yr = int(parts[2])
                        if yr >= 2018 or (yr >= 18 and yr <= 30):
                            if operator_date == "NOT FOUND":
                                operator_date = d
                            elif ussc_date == "NOT FOUND":
                                ussc_date = d
                    except:
                        pass

    return operator_date, ussc_name, ussc_date

# ─────────────────────────────────────────────
# Process single PDF
# ─────────────────────────────────────────────
def process_pdf_bytes(file_bytes, filename):
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
            full_text = '\n'.join(pages_text)
            op_date, ussc_name, ussc_date = extract_signature_fields(pages_text)
            contract_type = detect_contract_type(pages_text, full_text)
            return {
                "File": filename,
                "Director / Entity Name": extract_header_name(pages_text),
                "Contract Type": contract_type,
                "Date Operator Signed": op_date,
                "USSC Signer Name": ussc_name,
                "Date USSC Signed": ussc_date,
                "Contract Expiration Date": extract_expiration_date(full_text),
            }
    except Exception as e:
        return {
            "File": filename,
            "Director / Entity Name": f"ERROR: {e}",
            "Contract Type": "",
            "Date Operator Signed": "",
            "USSC Signer Name": "",
            "Date USSC Signed": "",
            "Contract Expiration Date": "",
        }

# ─────────────────────────────────────────────
# Build Excel
# ─────────────────────────────────────────────
def build_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Data"
    headers = [
        "File", "Director / Entity Name", "Contract Type",
        "Date Operator Signed", "USSC Signer Name",
        "Date USSC Signed", "Contract Expiration Date"
    ]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Contract type colors
    type_colors = {
        "Entity": "FFF3CD",               # soft amber
        "Individual with Entity Payment": "D1ECF1",  # soft blue
        "Individual": "D4EDDA",           # soft green
    }

    for row_num, record in enumerate(data, 2):
        contract_type = record.get("Contract Type", "")
        row_color = type_colors.get(contract_type, "FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=record.get(header, ""))
            cell.alignment = Alignment(vertical="center")
            cell.fill = PatternFill("solid", fgColor=row_color)

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────
st.markdown("# 📋 USSC Contract Extractor")
st.markdown('<p class="subtitle">Upload your Nike Sports Camp operator agreements and download a clean Excel summary in seconds.</p>', unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
    <strong>What this extracts:</strong> Director / Entity Name &nbsp;·&nbsp; Contract Type &nbsp;·&nbsp;
    Date Operator Signed &nbsp;·&nbsp; USSC Signer Name &nbsp;·&nbsp; Date USSC Signed &nbsp;·&nbsp; Contract Expiration Date
    <br><br>
    <strong>Supports:</strong> All contract years (2019–2026) &nbsp;·&nbsp; Individual &nbsp;·&nbsp; Entity &nbsp;·&nbsp; Individual with Entity Payment
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Upload contract PDFs",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if uploaded_files:
    st.markdown(f'<p class="file-count">✅ {len(uploaded_files)} file{"s" if len(uploaded_files) != 1 else ""} selected</p>', unsafe_allow_html=True)

    if st.button("Extract Contract Data"):
        results = []
        progress = st.progress(0)
        status = st.empty()

        for i, uploaded_file in enumerate(uploaded_files):
            status.markdown(f'<p class="file-count">Processing {i+1} of {len(uploaded_files)}: {uploaded_file.name}</p>', unsafe_allow_html=True)
            file_bytes = uploaded_file.read()
            result = process_pdf_bytes(file_bytes, uploaded_file.name)
            results.append(result)
            progress.progress((i + 1) / len(uploaded_files))

        status.empty()
        progress.empty()

        excel_bytes = build_excel(results)

        st.markdown(f"""
        <div class="result-box">
            🎉 Done! <strong>{len(results)} contract{"s" if len(results) != 1 else ""}</strong> processed successfully.
            <br><small>Rows are color-coded by contract type: 🟡 Entity &nbsp; 🔵 Individual with Entity Payment &nbsp; 🟢 Individual</small>
        </div>
        """, unsafe_allow_html=True)

        st.download_button(
            label="⬇️ Download Contract_Data.xlsx",
            data=excel_bytes,
            file_name="Contract_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.markdown("""
    <div style="border: 2px dashed #d1d5db; border-radius: 12px; padding: 2.5rem; text-align: center; color: #9ca3af; font-size: 0.9rem; margin-top: 0.5rem;">
        Drag and drop your PDF contracts here<br>or click <strong>Browse files</strong> above
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")
st.markdown('<p style="text-align:center; color:#9ca3af; font-size:0.8rem;">US Sports Camps · Nike Sports Camps · Contract Data Tool</p>', unsafe_allow_html=True)
